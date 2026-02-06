import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.chart import PieChart, LineChart, Reference
from openpyxl.styles import numbers

# Load Data 
file_paths = [
    "/Users/brandon/Downloads/Chase9149_Activity20250101_20251227_20251228.CSV"
    # add more files here as needed, make sure to add comma if you do
]

output_path = "/Users/brandon/Downloads/chase_spending_analysis.xlsx"

# Read and combine all CSVs
df_list = []

for path in file_paths:
    temp_df = pd.read_csv(path)
    df_list.append(temp_df)

df = pd.concat(df_list, ignore_index=True)

# Normalize column names
df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")

df["transaction_date"] = pd.to_datetime(
    df["transaction_date"],
    errors="coerce"
)

# Drop rows where date failed
df = df.dropna(subset=["transaction_date"])
# Remove rows where description contains "payment"
df = df[~df["description"].str.contains("payment", case=False, na=False)]

# Multiply all amounts by -1
df["amount"] = df["amount"] * -1


# Transaction Table
transaction_table = (
    df[["transaction_date", "description", "amount", "category"]]
    .sort_values("transaction_date")
)
transaction_table["transaction_date"] = (
    transaction_table["transaction_date"]
    .dt.strftime("%Y-%m")
)

# Monthly + Full Year Category Totals
df["month_str"] = df["transaction_date"].dt.strftime("%Y-%m")

monthly_category_totals = {}

for month, month_df in df.groupby("month_str"):
    totals = (
        month_df.groupby("category")["amount"]
        .sum()
        .sort_values(ascending=False)
        .reset_index()
        .rename(columns={"amount": "total_amount"})
    )

    # Add grand total row
    totals.loc[len(totals)] = [
        "GRAND TOTAL",
        totals["total_amount"].sum()
    ]

    monthly_category_totals[month] = totals

full_year_category_totals = (
    df.groupby("category")["amount"]
    .sum()
    .sort_values(ascending=False)
    .reset_index()
    .rename(columns={"amount": "total_amount"})
)

# Add grand total row
full_year_category_totals.loc[len(full_year_category_totals)] = [
    "GRAND TOTAL",
    full_year_category_totals["total_amount"].sum()
]


# Recurring Expenses

df["month"] = df["transaction_date"].dt.to_period("M")

recurring = []

for vendor, vendor_df in df.groupby("description"):
    if len(vendor_df) < 3:
        continue

    avg_amount = vendor_df["amount"].mean()
    amount_std = vendor_df["amount"].std()
    occurrences = len(vendor_df)
    months = vendor_df["month"].nunique()

    if amount_std <= 5 and months >= 3:
        recurring.append({
            "vendor": vendor,
            "average_amount": round(avg_amount, 2),
            "total_amount": round(avg_amount * occurrences, 2),
            "occurrences": occurrences,
            "months_active": months
        })

recurring_df = (
    pd.DataFrame(recurring)
    .sort_values(by="total_amount", ascending=False)
)

# Monthly Spending Forecast

# Monthly total spending
monthly_spend = (
    df.groupby("month_str")["amount"]
    .sum()
    .reset_index()
    .sort_values("month_str")
)

# Convert month index to numbers for trend line
monthly_spend["month_index"] = range(len(monthly_spend))

# Simple linear trend (y = mx + b)
x = monthly_spend["month_index"]
y = monthly_spend["amount"]

slope, intercept = np.polyfit(x, y, 1)

# Forecast next 12 months
forecast_months = 12
last_index = x.iloc[-1]

forecast_data = []

for i in range(1, forecast_months + 1):
    forecast_amount = slope * (last_index + i) + intercept
    forecast_data.append(forecast_amount)

# Build forecast dataframe
forecast_df = pd.DataFrame({
    "month": pd.date_range(
        start=pd.to_datetime(monthly_spend["month_str"].iloc[-1]) + pd.offsets.MonthBegin(1),
        periods=forecast_months,
        freq="MS"
    ).strftime("%Y-%m"),
    "predicted_spending": forecast_data
})

# Export to Excel
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    transaction_table.to_excel(writer, sheet_name="Transactions", index=False)
    recurring_df.to_excel(writer, sheet_name="Recurring Expenses", index=False)
    forecast_df.to_excel(writer, sheet_name="Spending Forecast", index=False)

    full_year_category_totals.to_excel(
        writer,
        sheet_name="Categories Full Year",
        index=False
    )

    for month in sorted(monthly_category_totals.keys()):
        monthly_category_totals[month].to_excel(
            writer,
            sheet_name=f"Categories {month}",
            index=False
        )
        
# Add Pie Charts to Category Sheets
wb = load_workbook(output_path)

def add_pie_chart(ws, title):
    max_row = ws.max_row - 1  # EXCLUDE GRAND TOTAL ROW

    if max_row < 2:
        return  # safety check

    pie = PieChart()
    pie.title = title

    data = Reference(ws, min_col=2, min_row=1, max_row=max_row)
    labels = Reference(ws, min_col=1, min_row=2, max_row=max_row)

    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)

    ws.add_chart(pie, "E2")

# Full year chart
add_pie_chart(wb["Categories Full Year"], "Spending by Category (Full Year)")

# Monthly charts
for sheet_name in wb.sheetnames:
    if sheet_name.startswith("Categories ") and sheet_name != "Categories Full Year":
        add_pie_chart(
            wb[sheet_name],
            f"Spending by Category ({sheet_name.replace('Categories ', '')})"
        )

# Add Line Chart to Forecast Sheet

forecast_ws = wb["Spending Forecast"]

line_chart = LineChart()
line_chart.title = "Projected Monthly Spending"
line_chart.y_axis.title = "Amount ($)"
line_chart.x_axis.title = "Month"

# Data range
data = Reference(
    forecast_ws,
    min_col=2,
    min_row=1,
    max_row=forecast_ws.max_row
)

# Month labels
categories = Reference(
    forecast_ws,
    min_col=1,
    min_row=2,
    max_row=forecast_ws.max_row
)

line_chart.add_data(data, titles_from_data=True)
line_chart.set_categories(categories)

# Set Y-axis bounds
forecast_values = [
    forecast_ws.cell(row=row, column=2).value
    for row in range(2, forecast_ws.max_row + 1)
    if isinstance(forecast_ws.cell(row=row, column=2).value, (int, float))
]

if forecast_values:
    line_chart.y_axis.scaling.min = min(forecast_values) - 100
    line_chart.y_axis.scaling.max = max(forecast_values) + 100

# Position chart
forecast_ws.add_chart(line_chart, "D2")

# Format money columns

money_format = '"$"#,##0.00'

for ws in wb.worksheets:
    # Read header row to find money columns
    headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    for col_name in ["amount", "total_amount", "average_amount"]:
        if col_name in headers:
            col_idx = headers[col_name]

            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = money_format

# Auto-fit column widths

for ws in wb.worksheets:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        # Add padding
        ws.column_dimensions[column_letter].width = max_length + 2

wb.save(output_path)

print("Excel file created with pie charts: {output_path}")
